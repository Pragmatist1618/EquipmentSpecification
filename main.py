import tkinter as tk
from tkinter import filedialog
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side


# Функция для чтения данных из Excel файлов в выбранной папке
def read_excel_files(folder_path):
    all_data = []
    for file in os.listdir(folder_path):
        if file.endswith('.xls'):
            file_path = os.path.join(folder_path, file)
            data = pd.read_excel(file_path, header=1)
            all_data.append((file[:-4], data))
    return all_data


# Функция для создания сводной таблицы
def create_pivot_table(data):
    if not data:
        print("No data found. Exiting...")
        return None

    dfs = []

    # Создание временных DataFrame для каждого файла и добавление их в список
    for file_name, df in data:
        col_name = file_name.split(' ')[2]
        try:
            temp_df = df[['Описание', 'Номер для заказа', 'Общ. количество (число штук)']].copy()
            temp_df.rename(columns={'Общ. количество (число штук)': col_name}, inplace=True)
            dfs.append(temp_df)
        except:
            print(f'Warning: "Описание" column not found in {col_name}. Skipping...')

    # Объединение данных из всех файлов по столбцу "Описание"
    pivot_table = pd.concat([df.set_index(['Описание', 'Номер для заказа']) for df in dfs], axis=1, join='outer')

    # Сортировка столбцов с количеством по алфавиту, учитывая числовые значения после дефиса
    pivot_table = pivot_table.reindex(
        sorted(pivot_table.columns, key=lambda x: (x.split('-')[0], int(x.split('-')[1]))), axis=1)

    # Сброс индекса
    pivot_table.reset_index(inplace=True)

    return pivot_table


# Функция для генерации сводной таблицы и сохранения в файле
def generate_pivot_table():
    # Запрос папки с файлами Excel
    folder_path = filedialog.askdirectory()

    # Чтение данных из файлов Excel в выбранной папке
    all_data = read_excel_files(folder_path)

    # Создание сводной таблицы
    pivot_table = create_pivot_table(all_data)

    # Сохранение сводной таблицы в новом файле Excel
    output_file = os.path.join(folder_path, 'Equipment Specification.xlsx')
    pivot_table.to_excel(output_file, index=False)

    # Открытие созданного файла для добавления форматирования
    wb = load_workbook(output_file)
    ws = wb.active

    # Выравнивание всех элементов кроме первого столбца по середине
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=ws.max_column, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Установка ширины столбцов
    ws.column_dimensions['A'].width = 70
    ws.column_dimensions['B'].width = 40

    # Форматирование первого столбца для переноса текста
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='center')

    # Добавление сетки
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # Сохранение изменений
    wb.save(output_file)


# Функция для создания графического интерфейса
def main():
    root = tk.Tk()
    root.title("Equipment Specification Generator")
    root.geometry("350x150")
    root.resizable(False, False)

    # Метка для инструкции
    label = tk.Label(root, text="Выберите папку с файлами Excel:")
    label.pack(pady=10)

    # Кнопка для выбора папки
    button_select = tk.Button(root, text="Выбрать папку", command=generate_pivot_table)
    button_select.pack(pady=5)

    # Кнопка для закрытия окна
    button_confirm = tk.Button(root, text="Закрыть", command=root.destroy)
    button_confirm.pack(pady=5)

    root.mainloop()


# Запуск приложения
if __name__ == "__main__":
    main()
